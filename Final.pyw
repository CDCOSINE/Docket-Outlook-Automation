# -*- coding: utf-8 -*-
"""
Created on Wed Apr 22 12:02:58 2020

@author: Jatin Sehgal-13 , jatinnitkkr@gmail.com , 7404060420
"""



### Necessary Libraries
import os
import win32com.client
#import dateutil.parser as dparser
import openpyxl
import datetime
import pandas as pd
import PySimpleGUI as sg
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font


##Path to desktop of any Machine/User
desktop = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')



## Email Read Screen
sg.theme('DarkAmber')   
layout1 = [[sg.Text('Enter Email/Name to look'),sg.InputText()],
            [sg.Text('On the next window Select Keywords csv File')],
            [sg.Button('Ok')] ]

window = sg.Window('Docket-Outlook-Automation',layout1)
a,b = window.read()
emaill = str(b[0])
window.close()
add = emaill





### Keywords CSV Dialog Box
root = tk.Tk()
root.withdraw()
filr_path = filedialog.askopenfilename(title='Select Keyword CSV File to choose for Keywords',filetypes = (("CSV Files","*.csv"),))





## List of Keywords from CSV
newkeyss = pd.read_csv(filr_path)
newkeyss = list(newkeyss['Keywords'])
print('abc',newkeyss)






## Start Date Calendar Screen
layout = [  [sg.Text('Choose Start Date')],
            [sg.In(key='-CAL-', enable_events=True, visible=False), sg.CalendarButton('Select Start Date', target='-CAL-', pad=None, font=('MS Sans Serif', 11, 'bold'),
                 key='_CALENDAR_', format=('%d %m, %Y'))]
           ]

window = sg.Window('Start Date', layout,size=(250,150))

event, values = window.read()
print(event,values)
window.close()






### Converting start date to DT OBJECT
svalues = values
splitval = svalues['-CAL-'].split(' ')
yearrr = int(splitval[2])
monthhh = int(splitval[1].replace(',',''))
dateee = int(splitval[0])
sjk = datetime.date(yearrr,monthhh,dateee)
start = sjk





##End Date Calendar Screen
layout = [  [sg.Text('Choose End Date')],
            [sg.In(key='-CAL-', enable_events=True, visible=False), sg.CalendarButton('Select End Date', target='-CAL-', pad=None, font=('MS Sans Serif', 11, 'bold'),
                 key='_CALENDAR_', format=('%d %m, %Y'))],
            [sg.Exit()]]

window = sg.Window('End Date', layout,size=(250,150))
event, values = window.read()
window.close()






## Convertng end date to DT object
evalues = values
splitval = evalues['-CAL-'].split(' ')
yearrr = int(splitval[2])
monthhh = int(splitval[1].replace(',',''))
dateee = int(splitval[0])
ejk = datetime.date(yearrr,monthhh,dateee)
end = ejk





### Calling Outlook MAPI
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")





### Commma separated Keywords
stringk = ' , '.join(newkeyss)


key = newkeyss


### INBOX FOLDER
inbox = outlook.GetDefaultFolder(6)





# TODAY DATE
dt = str(datetime.date.today())





####Creating Excel
wb = openpyxl.Workbook()
ws =  wb.active
ws.title = "Sheet1"
wb["Sheet1"]['A'+str(1)] = "Email Scrapped: " + add
wb["Sheet1"]['B'+str(1)] = "Key Searched " + stringk
wb["Sheet1"]['A'+str(2)] = "Keyword Matched"
wb["Sheet1"]['B'+str(2)] = 'Docket Update'
wb["Sheet1"]['C'+str(2)] = 'Case Name'
wb["Sheet1"]['D'+str(2)] = 'Case Number'
wb['Sheet1']['E'+str(2)] = 'Date of Email'
wb["Sheet1"]['F'+str(2)] = 'Email Title'
wb["Sheet1"]['G'+str(2)] = 'Campaign Name'
wb["Sheet1"]['H'+str(2)] = 'Campaign Link'
wb["Sheet1"]['I'+str(2)] = 'Docket Link'
wb.save(desktop+"/" +dt+stringk+add+ ".xlsx")




### ALL MESSAGES
messages = inbox.Items
found = ""
length=len(messages)






### Renaming Excel
work = openpyxl.load_workbook(desktop+"/" +dt+stringk+add+ ".xlsx")

sheet = work["Sheet1"]
k = 3





### Frame Progress Bar
outputwin = [
    [sg.Output(size=(78,20))]
]
progresslayout = [[sg.Text('Processing File')],
          [sg.ProgressBar(length, orientation='h', size=(20, 20), key='progressbar')],
          [sg.Frame('Output', layout = outputwin)],
          [sg.Button('Ok')],
          [sg.Cancel()]]
window = sg.Window('Processing File', progresslayout)
progress_bar = window['progressbar']




casenum='NA'
casename='NA'



## Delete Link from Docket Update


def delete(somestring):
    out = ''
    ind=1
    for letter in somestring:
#        print('l:  ',letter)
        if (letter=='<' or ind==0) and (letter!='>'):
#            print('c1')
            ind=0
            continue
        elif letter=='>':
#            print('c2')
            ind=1
            continue
        else:
#            print('c3')
            out = out+letter
    return out
    





#Backward Loop(RECENT FIRST)

for i in range(length-1,-1,-1):
    proc = length-i
    event, values = window.read(timeout=10)
    if event == 'Cancel'  or event == sg.WIN_CLOSED:
        break
    progress_bar.UpdateBar(proc + 1)
    if messages[i].Class == 43 and messages[i].senton.date()>=start and messages[i].senton.date()<=end:
        try:
            a=messages[i].Sender.GetExchangeUser().PrimarySmtpAddress
        except:
            a=messages[i].SenderEmailAddress
        if add.lower() in a.lower():
                liness = messages[i].body.lower().split("\n")
                lastcamp = ""
                lastlink = ""
                for c,lin in enumerate(liness):
                    if "campaign:" in str(lin).lower():
                        casenum = 'NA'
                        lis = str(lin).split("<")
                        lastcamp =  lis[0].split(": ")[1]
                        casename = str(liness[c+4]).split('<')[0]
                        print('')
                        print('New Case: ',lastcamp,' -- ',casename)
                        print('')
                    if casenum == '' and '-cv-' in str(lin):
                        casenum = str(lin).lower().split(' ')[0]
                    if casenum=='' and 'pgr' in str(lin):
                        casenum = str(lin).split(' ')[0].upper()
                    if casename == str(liness[c]).split('<')[0]:
                        casenum=''
                    for keyyy in key:
                        keyyy = keyyy.lower()
                        if keyyy == 'so':
                            lin = lin.upper()
                            keyyy = 'SO'
                        if keyyy == 'po':
                            lin = lin.upper()
                            keyyy = 'PO'
                        donot = 'parties terminated'
                        if keyyy in lin.lower() and (donot not in str(lin)):
                            print('')
                            print('(Key matched = {}) == in line ==  {}'.format(keyyy,lin))
                            print('')
                            found = str(lin)
                            f2 = found
                            found = delete(found)
                            sheet['A'+str(k)] = keyyy
                            sheet['G'+str(k)] = lastcamp
                            sheet['C'+str(k)] = casename
                            sheet['D'+str(k)] = casenum
                            sheet['E'+str(k)] = messages[i].senton.date()
                            sheet['B'+str(k)] = found
                            try:
                                sheet['H'+str(k)] = lis[1].split(">")[0]
                            except:
                                print("...",end=' ')
                            try:
                                l1 = f2.split("<")[1].split(">")[0]
                                sheet['I'+str(k)] = l1
                            except:
                                print('...',end=' ')
                            sheet['F'+str(k)] = messages[i].subject
                            k = k+1
print('------------------------------------OVER-----------------------------------')

                      
aa= window.read()
window.close()



### Resizing columns
for column_cells in sheet.columns:
    new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
    if (new_column_letter == 'A'):
        sheet.column_dimensions[new_column_letter].width = 20
        continue
    if (new_column_letter == 'D'):
        sheet.column_dimensions[new_column_letter].width = 16
        continue
    if (new_column_letter == 'E'):
        sheet.column_dimensions[new_column_letter].width = 14
        continue
    sheet.column_dimensions[new_column_letter].width = 32
    


#Path of Output File
print(desktop+"/" +dt+stringk+add+ ".xlsx")





#Bold Header Names
sheet['A2'].font = Font(bold=True)
sheet['B2'].font = Font(bold=True)
sheet['C2'].font = Font(bold=True)
sheet['D2'].font = Font(bold=True)
sheet['E2'].font = Font(bold=True)
sheet['F2'].font = Font(bold=True)
sheet['G2'].font = Font(bold=True)
sheet['H2'].font = Font(bold=True)
sheet['I2'].font = Font(bold=True)
sheet['J2'].font = Font(bold=True)


##Save File
work.save(desktop+"/" +dt+stringk+add+ ".xlsx")







### Screen Process Over
sg.theme('DarkAmber')   
layoutt = [[sg.Text('Process Over')],
            [sg.Text('Get your file at:{}'.format(desktop+"/" ))],
            [sg.Button('Ok')]]
window = sg.Window('Docket-Outlook-Automation',layoutt,size=(300,110))


a = window.read()
window.close()
